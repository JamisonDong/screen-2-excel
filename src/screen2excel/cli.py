"""命令行入口：screen2excel <图片...> -o out.xlsx"""

from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook

from .excel import write_table_html
from .ocr import image_to_table_htmls


def _clipboard_to_png() -> str:
    """把 macOS 剪贴板中的图片导出为临时 PNG 文件。"""
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.close()
    script = (
        "set pngData to the clipboard as «class PNGf»\n"
        f"set fp to open for access POSIX file \"{tmp.name}\" with write permission\n"
        "write pngData to fp\n"
        "close access fp"
    )
    r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if r.returncode != 0:
        raise SystemExit("读取剪贴板失败（剪贴板里没有图片？）: " + r.stderr.strip())
    return tmp.name


def _sheet_name(path: str, used: set[str]) -> str:
    name = Path(path).stem[:28] or "table"
    base, i = name, 2
    while name in used:
        name = f"{base}_{i}"
        i += 1
    used.add(name)
    return name


def main() -> None:
    p = argparse.ArgumentParser(
        prog="screen2excel",
        description="把表格截图转成 Excel（本地 PaddleOCR PP-StructureV3，无需联网/大模型）",
    )
    p.add_argument("images", nargs="*", help="截图文件路径（可多张，每张一个 sheet）")
    p.add_argument("--clipboard", action="store_true", help="从 macOS 剪贴板读取截图")
    p.add_argument("-o", "--output", required=True, help="输出 xlsx 路径")
    args = p.parse_args()

    images = list(args.images)
    if args.clipboard:
        images.append(_clipboard_to_png())
    if not images:
        p.error("请提供截图路径，或使用 --clipboard")

    for img in images:
        if not Path(img).exists():
            raise SystemExit(f"文件不存在: {img}")

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for img in images:
        htmls = image_to_table_htmls(img)
        if not htmls:
            print(f"警告: {img} 中未识别到表格，跳过", file=sys.stderr)
            continue
        ws = wb.create_sheet(_sheet_name(img, used))
        row_offset = 1
        for html in htmls:
            n = write_table_html(ws, html, start_row=row_offset)
            row_offset += n + 1  # 多个表格之间空一行
        print(f"✓ {img} → sheet「{ws.title}」（{len(htmls)} 个表格）")

    if not wb.sheetnames:
        raise SystemExit("所有图片都未识别到表格")
    wb.save(args.output)
    print(f"已保存: {args.output}")


if __name__ == "__main__":
    main()
