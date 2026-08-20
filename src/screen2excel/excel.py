"""HTML 表格（含 colspan/rowspan）→ openpyxl 工作表。"""

from __future__ import annotations

import re
from html.parser import HTMLParser

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_NUMBER_RE = re.compile(r"^-?\d[\d,]*\.?\d*$")


class _TableParser(HTMLParser):
    """把 <table> 解析成网格：grid[r][c] = (text, rowspan, colspan)。"""

    def __init__(self) -> None:
        super().__init__()
        self.grid: list[list[tuple[str, int, int]]] = []
        self._row: list[tuple[str, int, int]] | None = None
        self._cell: list[str] | None = None
        self._rs = 1
        self._cs = 1

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            a = dict(attrs)
            self._rs = int(a.get("rowspan") or 1)
            self._cs = int(a.get("colspan") or 1)
            self._cell = []

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            text = " ".join("".join(self._cell).split())
            self._row.append((text, self._rs, self._cs))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if self._row:
                self.grid.append(self._row)
            self._row = None

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)


def _to_number(text: str):
    """像数字的字符串转成 int/float，否则原样返回。"""
    if not _NUMBER_RE.match(text):
        return text
    s = text.replace(",", "")
    try:
        return int(s) if "." not in s else float(s)
    except ValueError:
        return text


def write_table_html(ws, html: str, start_row: int = 1) -> int:
    """把一段 HTML 表格从 start_row 开始写进工作表，返回占用行数。"""
    off = start_row - 1
    parser = _TableParser()
    parser.feed(html)
    grid = parser.grid

    # 裁掉所有行（含表头）末格都为空的尾部列（表格识别常会多出一列空列）
    def _trailing_col_empty() -> bool:
        rows_with_cells = [row for row in grid if row]
        return bool(rows_with_cells) and all(not row[-1][0] for row in rows_with_cells)

    while _trailing_col_empty():
        for row in grid:
            if not row:
                continue
            text, rs, cs = row[-1]
            if cs > 1:
                row[-1] = (text, rs, cs - 1)
            else:
                row.pop()

    # 展开 rowspan/colspan 到平面网格，记录合并区域
    ncols = max((sum(cs for _, _, cs in row) for row in grid), default=0)
    cells: dict[tuple[int, int], str] = {}
    merges: list[tuple[int, int, int, int]] = []
    occupied: set[tuple[int, int]] = set()
    for r, row in enumerate(grid):
        c = 0
        for text, rs, cs in row:
            while (r, c) in occupied:
                c += 1
            cells[(r, c)] = text
            if rs > 1 or cs > 1:
                merges.append((r, c, r + rs - 1, c + cs - 1))
                for dr in range(rs):
                    for dc in range(cs):
                        occupied.add((r + dr, c + dc))
            c += cs

    for (r, c), text in cells.items():
        ws.cell(row=r + 1 + off, column=c + 1, value=_to_number(text))
    for r1, c1, r2, c2 in merges:
        ws.merge_cells(
            start_row=r1 + 1 + off,
            start_column=c1 + 1,
            end_row=r2 + 1 + off,
            end_column=c2 + 1,
        )

    # 表头加粗
    if grid:
        for c in range(ncols):
            cell = ws.cell(row=1 + off, column=c + 1)
            cell.font = Font(bold=True)

    # 列宽自适应（按字符宽度估算，中文按 2 计）
    for c in range(ncols):
        width = 10
        for r in range(len(grid)):
            v = ws.cell(row=r + 1 + off, column=c + 1).value
            if v is not None:
                w = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                width = max(width, min(w + 2, 60))
        cur = ws.column_dimensions[get_column_letter(c + 1)].width or 0
        ws.column_dimensions[get_column_letter(c + 1)].width = max(cur, width)

    return len(grid)
